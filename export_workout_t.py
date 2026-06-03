import os
import datetime
import json
import openpyxl

excel_path = r"C:\Users\visua\Google Drive\Job Personal Trainer\Contabilita\Backup\MoneyManager 23 feb 2026.xlsx"

if not os.path.exists(excel_path):
    print(f"Error: {excel_path} not found.")
    exit(1)

wb = openpyxl.load_workbook(excel_path, data_only=True)
if "WORKOUT_T" not in wb.sheetnames:
    print("Error: WORKOUT_T sheet not found in the workbook.")
    exit(1)

sheet = wb["WORKOUT_T"]
rows = list(sheet.iter_rows(values_only=True))

headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
print("Headers found in WORKOUT_T:", headers)

records = []
for row_idx, row in enumerate(rows[1:], start=2):
    # Check if the row has any non-empty data
    if all(v is None for v in row):
        continue
        
    record = {}
    for col_idx, val in enumerate(row):
        header = headers[col_idx]
        
        # Format date values to DD/MM/YYYY string
        if isinstance(val, (datetime.datetime, datetime.date)):
            val = val.strftime("%d/%m/%Y")
        
        # Handle floats/numbers to avoid .0 for IDs and counts
        elif isinstance(val, float):
            if val.is_integer():
                val = int(val)
                
        # Handle None values
        if val is None:
            val = ""
            
        record[header] = val
        
    # Ensure ID_cliente and num_scheda are treated as strings to avoid format discrepancy
    if "ID_cliente" in record:
        record["ID_cliente"] = str(record["ID_cliente"]).strip()
    if "num_scheda" in record:
        record["num_scheda"] = str(record["num_scheda"]).strip()
        
    # Only keep records that have ID_cliente and num_scheda
    if record.get("ID_cliente") and record.get("num_scheda"):
        records.append(record)

print(f"Processed {len(records)} records from WORKOUT_T.")

with open("workout_t.json", "w", encoding="utf-8") as f:
    json.dump(records, f, indent=2, ensure_ascii=False)
print("Wrote workout_t.json successfully.")
